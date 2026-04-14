#!/usr/bin/env python3
"""Parse the TFP 2021 Online Supplement xlsx into tfp_prices.csv.

USDA publishes per-FNDDS-food prices at $/100g as-consumed, in June 2021
dollars. This is a much richer source than the PDF's 58-category table —
3,072 individual foods directly indexed by FNDDS Food Code.

Output columns:
  fndds_code          int       FNDDS Food Code (USDA's per-food ID)
  tfp_category        str       TFP modeling category
  pricing_method      int       PPPT pricing method code
  price_per_100g_2021 float     as-consumed, June 2021 dollars

Download:
  https://www.fns.usda.gov/cnpp/thrifty-food-plan-2021 →
  "Supplement to Thrifty Food Plan, 2021" .xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHEET = "Online Supplement Data"
EXPECTED_HEADERS = (
    "FNDDS Food Code",
    "TFP Modeling Category",
    "PPPT Pricing Methodb",
    "June 2021 as-consumed price per 100gc",
)


def parse(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if EXPECTED_SHEET not in wb.sheetnames:
        raise SystemExit(f"Expected sheet '{EXPECTED_SHEET}' not found in {xlsx_path}")
    ws = wb[EXPECTED_SHEET]

    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    for i, expected in enumerate(EXPECTED_HEADERS):
        if headers[i] != expected:
            raise SystemExit(
                f"Column {i}: expected '{expected}', got '{headers[i]}'. "
                "Has USDA changed the xlsx layout?"
            )

    out = []
    anomalies = []
    for rn, row in enumerate(rows, start=2):
        code, cat, method, price = row[0], row[1], row[2], row[3]
        if code is None or price is None:
            continue
        try:
            price = float(price)
        except (TypeError, ValueError):
            anomalies.append(f"row {rn}: price not numeric ({price!r})")
            continue
        if not 0.001 <= price <= 30.0:
            # $30/100g = $135/lb; above this is almost certainly a parse error.
            # Raised from 5.00 after confirming seafood legitimately hits $5-25/100g.
            anomalies.append(f"row {rn}: price ${price:.4f}/100g out of sanity range (0.001-30.00)")
        out.append({
            "fndds_code": int(code),
            "tfp_category": cat or "",
            "pricing_method": int(method) if method is not None else 0,
            "price_per_100g_2021": price,
        })

    if anomalies:
        print(f"[warn] {len(anomalies)} anomalies flagged:", file=sys.stderr)
        for a in anomalies[:20]:
            print(f"  {a}", file=sys.stderr)
        if len(anomalies) > 20:
            print(f"  ... and {len(anomalies) - 20} more", file=sys.stderr)

    return out


def write_csv(rows: list[dict], out_path: Path) -> None:
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=["fndds_code", "tfp_category", "pricing_method", "price_per_100g_2021"]
        )
        w.writeheader()
        w.writerows(rows)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default="data/raw/TFP_2021_Online_Supplement.xlsx")
    p.add_argument("--output", default="data/tfp_prices.csv")
    args = p.parse_args()

    rows = parse(Path(args.xlsx))
    write_csv(rows, Path(args.output))
    categories = {r["tfp_category"] for r in rows}
    print(
        f"wrote {len(rows)} rows ({len(categories)} distinct TFP categories) "
        f"to {args.output}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
